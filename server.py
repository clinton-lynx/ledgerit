"""Ledgerit's local HTTP server — the real backend behind worker/web/.

Two jobs:
  - serve the static app (worker/web/index.html), same origin only
  - a small JSON API wired directly to the real pipeline: cleaner.clean(),
    analyst.answer(), explain.classify(), explain.explain(). No mock data,
    no separate demo layer — this calls the exact same functions main.py
    calls.

Stdlib only (http.server + json). No Flask/FastAPI: the target machine has
an 8 GB total / 7 GB hard-limit budget the model already dominates, and a
web framework isn't worth the memory here for three JSON endpoints.
"""
from __future__ import annotations

import base64
import binascii
import csv
import http.server
import io
import json
import re
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WORKER_DIR = Path(__file__).resolve().parent
WEB_DIR = WORKER_DIR / "web"
SAMPLE_CSV = WORKER_DIR / "data" / "sales_raw.csv"

sys.path.insert(0, str(WORKER_DIR))

import pandas as pd  # noqa: E402

from analyst import HANDLERS, answer, build_index  # noqa: E402
from cleaner import EmptyDatasetError, MissingColumnsError, clean, infer_column_mapping  # noqa: E402
from explain import QuestionTooLongError, classify, explain  # noqa: E402
from explain import _llm as _load_model  # noqa: E402


# --------------------------------------------------------------------------
# Reading whatever file the browser sent up.
#
# The upload always arrives as base64 bytes (FileReader.readAsDataURL on
# the frontend, for both CSV and XLSX — see web/build.py) rather than
# CSV-as-text. That's deliberate: decoding text in the browser before
# sending it means committing to an encoding (JS defaults to UTF-8) before
# we ever get a look at the bytes, and a shop owner's CSV exported from
# Excel on Windows is routinely Windows-1252 or UTF-8-with-BOM, not UTF-8.
# Once JS has mis-decoded those bytes, the original bytes are gone — there
# is nothing left to detect an encoding from. Sending raw bytes and doing
# the decoding here, where we can try several encodings against the actual
# bytes, is the only way to recover the file as the owner's spreadsheet
# program actually wrote it.
# --------------------------------------------------------------------------

def _decode_csv_bytes(raw: bytes) -> str:
    """Try encodings in order of how likely they are to be both correct
    and detectable, not just "won't raise" — latin-1 never raises (it maps
    every byte 0-255 to a code point) so it's listed last, as the true
    last resort, not a first guess that happens to always succeed."""
    # UTF-16 has an unambiguous byte-order-mark; check for it explicitly
    # rather than letting it fall into the same try/except as the others,
    # since decoding UTF-16 bytes as a single-byte encoding "succeeds" with
    # garbage instead of raising, which would hide the real encoding.
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return raw.decode("utf-16")
    for enc in ("utf-8-sig", "windows-1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")  # unreachable: latin-1 never raises


def _sniff_delimiter(text: str) -> str:
    """Comma is the overwhelming default, but a semicolon-delimited export
    (common from Excel set to a European/Nigerian locale, where comma is
    the decimal separator) or a tab-delimited one otherwise loads as a
    single mangled column with no error — csv.Sniffer catches that instead
    of pandas silently "succeeding" at reading the file wrong."""
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","  # sniffer gives up on very short/single-column samples; comma is the safe default


def _snake(name: str) -> str:
    name = name.strip().lower()
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"[^a-z0-9]+", "_", name)
    return name.strip("_")


def _slugify(text: str) -> str:
    slug = _snake(text)
    slug = re.sub(r"_+", "-", slug).strip("-")
    return slug or "ledgerit"


def _today_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _export_root() -> Path:
    root = Path.home() / "Desktop" / "Ledgerit"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _wrap_lines(text: str, width: int = 72) -> list[str]:
    lines: list[str] = []
    for paragraph in str(text).splitlines() or [""]:
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            if len(current) + 1 + len(word) <= width:
                current += " " + word
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines or [""]


def _format_leader(left: str, right: str, width: int = 56) -> str:
    left = str(left)
    right = str(right)
    gap = max(2, width - len(left) - len(right))
    return f"{left}{'.' * gap}{right}"


def _render_pdf(lines: list[tuple[str, str, int]], title: str) -> bytes:
    page_w, page_h = 595, 842  # A4 in points
    margin_x = 40
    top = 792
    bottom = 44
    line_height = 14

    pages: list[list[tuple[str, str, int]]] = []
    current: list[tuple[str, str, int]] = []
    y = top
    for item in lines:
        if current and y < bottom:
            pages.append(current)
            current = []
            y = top
        current.append(item)
        y -= line_height
    if current:
        pages.append(current)
    if not pages:
        pages = [[("Courier", "", 10)]]

    obj_count = 4 + len(pages) * 2
    objects: list[bytes | None] = [None] * (obj_count + 1)
    page_kids: list[str] = []
    objects[1] = b"<< /Type /Catalog /Pages 2 0 R >>"
    objects[2] = b"<< /Type /Pages /Kids [] /Count 0 >>"
    objects[3] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>"
    objects[4] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier-Bold >>"

    for page_idx, page_lines in enumerate(pages):
        page_obj = 5 + page_idx * 2
        content_obj = page_obj + 1
        page_kids.append(f"{page_obj} 0 R")

        content_parts = [b"BT"]
        cur_font = None
        cur_size = None
        y = top
        for font_name, text, size in page_lines:
            font_id = "F1" if font_name == "Courier" else "F2"
            if font_name != cur_font or size != cur_size:
                content_parts.append(f"/{font_id} {size} Tf".encode("ascii"))
                cur_font = font_name
                cur_size = size
            content_parts.append(f"1 0 0 1 {margin_x} {y} Tm".encode("ascii"))
            content_parts.append(f"({_pdf_escape(text)}) Tj".encode("utf-8"))
            y -= line_height
        content_parts.append(b"ET")
        content = b"\n".join(content_parts)
        objects[page_obj] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_obj} 0 R >>".encode("ascii")
        )
        objects[content_obj] = (
            b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream"
        )

    objects[2] = f"<< /Type /Pages /Kids [{' '.join(page_kids)}] /Count {len(pages)} >>".encode("ascii")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0] * (len(objects))
    for idx in range(1, len(objects)):
        obj = objects[idx]
        if obj is None:
            continue
        offsets[idx] = len(pdf)
        pdf.extend(f"{idx} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_pos = len(pdf)
    pdf.extend(f"xref\n0 {len(objects)}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        (
            "trailer\n"
            f"<< /Size {len(objects)} /Root 1 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(pdf)


def _report_lines(filename: str, report: dict, title: str, body_lines: list[str]) -> bytes:
    lines: list[tuple[str, str, int]] = [
        ("Courier-Bold", title.upper(), 16),
        ("Courier", f"File: {filename}", 10),
        ("Courier", _format_leader("Rows in file", str(report.get("rows_in", ""))), 10),
        ("Courier", _format_leader("Rows kept", str(report.get("rows_out", ""))), 10),
    ]
    if report.get("duplicates_removed"):
        lines.append(("Courier", _format_leader("Duplicate rows removed", str(report["duplicates_removed"])), 10))
    if report.get("currency_stripped"):
        lines.append(("Courier", _format_leader("Currency text cleaned", str(report["currency_stripped"])), 10))
    if report.get("dates_failed"):
        lines.append(("Courier", _format_leader("Dates that could not be read", str(report["dates_failed"])), 10))
    for key, value in (report.get("blanks_filled") or {}).items():
        lines.append(("Courier", _format_leader(f"Missing {key.replace('_', ' ')} filled", str(value)), 10))
    lines.append(("Courier-Bold", "", 10))
    for body in body_lines:
        lines.append(("Courier", body, 10))
    return _render_pdf(lines, title)


def _finding_lines(question: str, finding, narration: str | None) -> list[tuple[str, str, int]]:
    lines: list[tuple[str, str, int]] = [
        ("Courier-Bold", "ANSWER", 16),
        ("Courier", f"Question: {question}", 10),
        ("Courier", f"Headline: {finding.headline}", 10),
        ("Courier-Bold", "", 10),
        ("Courier-Bold", "COMPUTED TABLE", 12),
    ]
    if finding.table is not None and not finding.table.empty:
        table = finding.table.to_dict(orient="records")
        label_col = list(table[0].keys())[0]
        value_col = "revenue" if "revenue" in table[0] else list(table[0].keys())[1]
        for row in table:
            label = str(row.get(label_col, ""))
            value = str(row.get(value_col, ""))
            lines.append(("Courier", _format_leader(label, value), 10))
    if finding.facts:
        lines.append(("Courier-Bold", "", 10))
        lines.append(("Courier-Bold", "FACTS", 12))
        for key, value in finding.facts.items():
            for idx, line in enumerate(_wrap_lines(str(value), 68)):
                prefix = f"{key.replace('_', ' ').title()}: " if idx == 0 else " " * (len(key) + 2)
                lines.append(("Courier", prefix + line, 10))
    if narration:
        lines.append(("Courier-Bold", "", 10))
        lines.append(("Courier-Bold", "NARRATION", 12))
        for line in _wrap_lines(narration, 74):
            lines.append(("Courier", line, 10))
    return lines


def _cleaning_lines(report: dict) -> list[str]:
    lines: list[str] = []
    if report.get("total_mismatches_count", 0):
        lines.append(f"Entries that don't add up: {report['total_mismatches_count']}")
        for item in report.get("total_mismatches") or []:
            lines.append(
                f"{item['order_id']}  recorded NGN {item['recorded_total']:,.0f}  "
                f"should be NGN {item['expected_total']:,.0f}"
            )
    else:
        lines.append("No arithmetic mismatches were found.")
    return lines


# Above this raw size, the memory cost of holding the previous dataset
# alongside a new one being validated stops being negligible against the
# 7 GB budget — measured up to ~2.92 GB peak RSS with the model warm, at
# 2M rows / 137 MB (docs/adversarial-testing-2026-08-07.md, #5). A typical
# shop's export is thousands to tens of thousands of rows — a few hundred
# KB to a few MB — where holding both is genuinely negligible, and where
# keeping the old dataset available as a fallback (in case the new upload
# turns out to be invalid) is worth far more than the memory it costs.
# Only above this backstop does that safety net get traded away for
# headroom; nothing this size was actually observed to risk the ceiling —
# it's a conservative line drawn well inside the smallest size that was.
_LARGE_UPLOAD_BYTES = 50 * 1024 * 1024  # ~50 MB raw (base64 length is checked pre-decode)


def _is_large_upload(body: dict) -> bool:
    # base64 expands bytes by ~4/3; comparing the encoded string's length
    # directly against a raw-byte threshold undercounts slightly, which
    # only makes this trigger a little earlier, not later.
    return len(body.get("file") or "") > _LARGE_UPLOAD_BYTES


class NeedsSheetSelectionError(Exception):
    """Raised when an xlsx workbook has more than one sheet. There is no
    safe way to guess which one holds the sales data — a "Summary" or
    "Pivot" tab ahead of the real data is common — so this always asks
    rather than silently reading whichever sheet openpyxl lists first."""

    def __init__(self, filename: str, sheets: list[dict]) -> None:
        self.filename = filename
        self.sheets = sheets
        super().__init__("multiple sheets — needs a choice")


class NeedsHeaderRowError(Exception):
    """Raised when the header-row detector can't tell, with a comfortable
    margin, which of the first several rows is the real header. Guessing
    wrong here is worse than a crash — see _detect_header_row."""

    def __init__(self, filename: str, sheet: str, candidates: list[dict]) -> None:
        self.filename = filename
        self.sheet = sheet
        self.candidates = candidates
        super().__init__("ambiguous header row — needs a choice")


def _describe_sheets(raw: bytes, sheet_names: list[str]) -> list[dict]:
    """One line of context per sheet — row/column counts and a peek at its
    first row — so the user can tell "Summary" from "RawSales" without
    opening the file themselves."""
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        out = []
        for name in sheet_names:
            ws = wb[name]
            preview: list[str] = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                preview = [str(v).strip() for v in row if v is not None and str(v).strip()][:6]
                break
            out.append({
                "name": name,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "preview": preview,
            })
        return out
    finally:
        wb.close()


def _row_profile(values: list) -> dict:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    width = len(values) or 1
    text_count = sum(1 for v in non_empty if isinstance(v, str))
    return {
        "non_empty": len(non_empty),
        "fill_ratio": len(non_empty) / width,
        "text_ratio": (text_count / len(non_empty)) if non_empty else 0.0,
        "distinct_ratio": (len({str(v).strip().lower() for v in non_empty}) / len(non_empty)) if non_empty else 0.0,
        "avg_len": (sum(len(str(v)) for v in non_empty) / len(non_empty)) if non_empty else 0.0,
    }


def _detect_header_row(raw: bytes, sheet: str, scan_rows: int = 15) -> dict:
    """Scores each of the first `scan_rows` rows on how header-like it is
    — well-filled, mostly text, distinct values, and noticeably more
    "texty" than the rows right after it (real headers are short labels;
    the data below them is a mix of numbers, dates and text). The row with
    the clear best score wins; if the best and second-best are too close,
    or nothing scores well, this is genuinely ambiguous and the caller
    should ask rather than guess — thresholds tuned against real fixtures
    (see docs/xlsx-header-detection notes), not picked in the abstract."""
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        max_col = ws.max_column or 1
        limit = min(scan_rows, ws.max_row or 1)
        rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=limit, max_col=max_col, values_only=True)]
    finally:
        wb.close()

    if not rows:
        return {"row": 0, "confident": True, "candidates": []}

    profiles = [_row_profile(r) for r in rows]
    scores = []
    for i, p in enumerate(profiles):
        if p["non_empty"] < 2:  # a near-blank row (title, spacer) can't be a header
            scores.append(-1.0)
            continue
        follow = profiles[i + 1:i + 4]
        follow_text = (sum(f["text_ratio"] for f in follow) / len(follow)) if follow else p["text_ratio"]
        score = (
            p["fill_ratio"] * 1.0
            + p["text_ratio"] * 1.5
            + p["distinct_ratio"] * 0.5
            - follow_text * 0.8
            - (0.3 if p["avg_len"] > 40 else 0.0)  # long strings read as prose/titles, not column labels
        )
        scores.append(score)

    ranked = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)
    best = ranked[0]
    second = scores[ranked[1]] if len(ranked) > 1 else -999.0
    confident = (scores[best] - second) >= 0.3 and scores[best] >= 1.2

    candidates = [
        {"row": i, "score": round(scores[i], 3), "preview": [str(v).strip() for v in rows[i] if v is not None][:6]}
        for i in ranked[:5] if scores[i] > -1.0
    ]
    return {"row": best, "confident": confident, "candidates": candidates}


def _read_xlsx(raw: bytes, filename: str, body: dict) -> pd.DataFrame:
    """Handles the two ambiguities plain pd.read_excel(...) papers over by
    always reading sheet 0 with header=0: which sheet has the data (asked
    explicitly, every time there's more than one — see
    NeedsSheetSelectionError) and which row is the header (auto-detected
    with a confidence gate — see _detect_header_row / NeedsHeaderRowError).
    `body` carries the user's answers back in on the next request: `sheet`
    and/or `header_row`. Both are re-derived from the raw bytes on every
    call rather than cached — consistent with the rest of this module,
    which never holds server-side state between the mapping preview and
    the confirm step."""
    try:
        sheet_names = pd.ExcelFile(io.BytesIO(raw), engine="openpyxl").sheet_names
    except Exception as e:
        raise ValueError(f"Couldn't read that as an Excel file: {e}")
    if not sheet_names:
        raise ValueError("That workbook doesn't have any sheets.")

    chosen_sheet = body.get("sheet")
    if len(sheet_names) > 1:
        if chosen_sheet not in sheet_names:
            raise NeedsSheetSelectionError(filename, _describe_sheets(raw, sheet_names))
        sheet = chosen_sheet
    else:
        sheet = sheet_names[0]

    header_row = body.get("header_row")
    if header_row is None:
        detected = _detect_header_row(raw, sheet)
        if not detected["confident"]:
            raise NeedsHeaderRowError(filename, sheet, detected["candidates"])
        header_row = detected["row"]
    else:
        try:
            header_row = int(header_row)
        except (TypeError, ValueError):
            raise ValueError("That header row selection wasn't valid — try again.")

    try:
        return pd.read_excel(io.BytesIO(raw), engine="openpyxl", sheet_name=sheet, header=header_row)
    except Exception as e:
        raise ValueError(f"Couldn't read that as an Excel file: {e}")


def _read_uploaded_dataframe(body: dict) -> tuple[pd.DataFrame, str]:
    """Returns (raw_dataframe, filename). Raises ValueError with a message
    fit to show the user directly if the upload can't be read at all
    (corrupt file, wrong format, undecodable bytes); raises
    NeedsSheetSelectionError / NeedsHeaderRowError when reading it requires
    an answer only the user can give."""
    # "no 'file' key at all" (the sample-load / replay path — see
    # loadFile(null, null) in web/build.py) and "'file' key present but its
    # value is falsy" are different situations and used to be handled
    # identically, because base64.b64encode(b"") == "" is falsy. A genuine
    # 0-byte upload silently loaded the bundled sample instead of erroring
    # (docs/adversarial-testing-2026-08-07.md, #1) — checking key presence
    # instead of truthiness tells the two apart.
    if "file" not in body:
        return pd.read_csv(SAMPLE_CSV), SAMPLE_CSV.name

    file_b64 = body.get("file") or ""
    filename = (body.get("filename") or "").strip()

    if not file_b64:
        raise ValueError("That file appears to be empty — nothing to load.")

    try:
        raw = base64.b64decode(file_b64, validate=True)
    except (binascii.Error, ValueError):
        raise ValueError("That upload didn't arrive intact — try again.")

    is_xlsx = filename.lower().endswith(".xlsx") or raw[:4] == b"PK\x03\x04"
    if is_xlsx:
        df = _read_xlsx(raw, filename or "upload.xlsx", body)
        out_name = filename or "upload.xlsx"
    else:
        try:
            text = _decode_csv_bytes(raw)
        except Exception as e:  # pragma: no cover - latin-1 fallback makes this unreachable in practice
            raise ValueError(f"Couldn't read that file's text encoding: {e}")

        delimiter = _sniff_delimiter(text)
        try:
            df = pd.read_csv(io.StringIO(text), sep=delimiter)
        except pd.errors.EmptyDataError:
            raise ValueError("That file is empty — nothing to read.")
        except Exception as e:
            raise ValueError(f"Couldn't read that as a CSV: {e}")
        out_name = filename or "upload.csv"

    # openpyxl hands back whatever type a header cell holds — int, float,
    # datetime — not just str, whenever a header row has a non-text cell
    # (a stray year, a mis-typed serial number). Every column-name string
    # operation downstream (snake-casing, mapping inference) assumes str;
    # coercing once here, at the single point every read path returns
    # through, is cheaper and more reliable than patching each call site.
    df.columns = [str(c).strip() for c in df.columns]
    return df, out_name


class _State:
    """What has to persist between requests: the cleaned frame and its BM25
    index. Guarded by a lock — ThreadingHTTPServer means two requests can
    arrive at once, and swapping `df` out from under a read isn't safe."""

    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.df = None
        self.index = None
        self.filename: str | None = None
        self.mapping: dict[str, str] | None = None


STATE = _State()

# llama_cpp's loaded model is one C++ context with internal, mutable
# generation state (its KV cache) — it is not safe to call into from two
# threads at once. ThreadingHTTPServer hands each request its own thread,
# and the frontend's /api/ask and /api/narrate calls can be in flight
# together, so every call that reaches the model (classify() or explain(),
# both funnel through explain._llm()) has to be serialised through this
# lock. Found by running the real UI end to end: firing both requests
# concurrently segfaulted the server (SIGSEGV) the first time a question
# was asked. Pandas-only work (answer()) never touches this lock.
MODEL_LOCK = threading.Lock()


def warm_model_in_background() -> None:
    """Load the model once, before the user ever asks anything — otherwise
    the first real question pays the ~15-50s model-load cost on top of
    generation time, which reads as the app being broken, not just slow."""
    def _warm():
        try:
            print("warming the model…", file=sys.stderr)
            with MODEL_LOCK:
                _load_model()
            print("model ready", file=sys.stderr)
        except Exception as e:  # pragma: no cover - best-effort warm-up
            print(f"model warm-up failed, will load on first request instead: {e}", file=sys.stderr)

    threading.Thread(target=_warm, daemon=True).start()


def _finding_to_json(finding, question: str, route: str | None) -> dict:
    """Finding -> JSON for the frontend.

    Deliberately excludes `finding.guidance`: that field is instruction
    text for the model — how to read the numbers so it doesn't invert them
    — not something a shop owner should ever see rendered as if it were
    part of their data.
    """
    table = None
    if finding.table is not None and not finding.table.empty:
        # round-trip through pandas' own JSON encoder rather than
        # DataFrame.to_dict(), which leaves numpy int64/float64 in the
        # values and those aren't JSON-serialisable by the stdlib encoder.
        table = json.loads(finding.table.to_json(orient="records"))
    return {
        "question": question,
        "route": route,
        "headline": finding.headline,
        "facts": finding.facts or {},
        "table": table,
    }


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(WEB_DIR), **kwargs)

    def guess_type(self, path):
        """SimpleHTTPRequestHandler's guessed Content-Type never includes a
        charset. Serve this for real instead of through the artifact
        preview and that's exactly the gap that showed up: the receipt's
        emoji, its em dash, its multiplication sign all rendered as
        mojibake, because nothing told the browser this file is UTF-8. The
        page's own <meta charset="utf-8"> is the belt; this header is the
        suspenders — the HTTP header wins if the two ever disagree."""
        ctype = super().guess_type(path)
        if ctype.startswith("text/") or ctype == "application/javascript":
            return ctype + "; charset=utf-8"
        return ctype

    def log_message(self, fmt, *args):
        sys.stderr.write("%s - %s\n" % (self.address_string(), fmt % args))

    # ---------------------------------------------------------------- #
    # JSON helpers
    # ---------------------------------------------------------------- #
    def _send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        try:
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError):
            pass

    def _read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", 0) or 0)
        raw = self.rfile.read(length) if length else b""
        return json.loads(raw.decode("utf-8")) if raw else {}

    # ---------------------------------------------------------------- #
    # routes
    # ---------------------------------------------------------------- #
    def do_POST(self):
        try:
            if self.path == "/api/load":
                return self._handle_load()
            if self.path == "/api/ask":
                return self._handle_ask()
            if self.path == "/api/narrate":
                return self._handle_narrate()
            if self.path == "/api/export":
                return self._handle_export()
            self._send_json({"error": "not found"}, 404)
        except Exception as e:
            # str(e) alone is fine for the ValueErrors handlers raise on
            # purpose (already written to be shown to the user), but a
            # genuinely unexpected exception — a bug, not bad input — used
            # to hand the user a raw Python message ("'int' object has no
            # attribute 'strip'") with nothing in the server log to debug
            # it from. The traceback goes to stderr where it's actually
            # useful; the user gets a message that doesn't assume they
            # know what an AttributeError is.
            sys.stderr.write(f"unhandled error on POST {self.path}:\n{traceback.format_exc()}")
            self._send_json({"error": f"Something went wrong on the server's end ({type(e).__name__}). Try again."}, 500)

    def _handle_load(self):
        """Real cleaner.clean() run, on either the bundled sample or a file
        the browser sent up (CSV or XLSX, as base64 bytes — see
        _read_uploaded_dataframe for why not text). Bad input (wrong
        format, undecodable bytes, missing columns, nothing left after
        cleaning) comes back as 400 with a message meant to be shown to the
        user directly, not a 500 with a raw traceback three modules away
        from where the file was read."""
        body = self._read_json()

        # Only a large upload (see _LARGE_UPLOAD_BYTES) frees the previous
        # dataset up front. Below that size, validate-then-swap is strictly
        # better: the old dataset stays available as a fallback the whole
        # time, and only gets replaced once the new one is known-good —
        # so a bad upload never costs the user their working data. That
        # safety net is only traded away above the size where holding both
        # stops being negligible against the memory budget.
        large_upload = _is_large_upload(body)
        if large_upload:
            with STATE.lock:
                STATE.df = None
                STATE.index = None

        try:
            raw_df, filename = _read_uploaded_dataframe(body)
        except NeedsSheetSelectionError as e:
            return self._send_json({
                "needs_sheet_selection": True,
                "filename": e.filename,
                "sheets": e.sheets,
            })
        except NeedsHeaderRowError as e:
            return self._send_json({
                "needs_header_row": True,
                "filename": e.filename,
                "sheet": e.sheet,
                "candidates": e.candidates,
            })
        except ValueError as e:
            return self._send_json({"error": str(e)}, 400)

        normalized = {_snake(c) for c in raw_df.columns}
        exact_required = all(req in normalized for req in ("date", "product", "qty", "unit_price", "total"))
        mapping_body = body.get("mapping") or None
        inferred = None

        # The base64 payload (body["file"]) can be well over 100 MB for a
        # large file and isn't needed again — drop it before the
        # clean()/build_index() work below rather than let it sit resident
        # in `body` for the rest of this method purely because `body` is
        # still in scope.
        body.clear()

        if mapping_body is None and not exact_required:
            inferred = infer_column_mapping(raw_df, remembered=STATE.mapping)
            if inferred["essential_missing"]:
                msg = "This file is missing a column Ledgerit needs: " + ", ".join(inferred["essential_missing"]) + "."
                return self._send_json({"error": msg}, 400)
            self._send_json({
                "needs_mapping": True,
                "filename": filename,
                "required": inferred["targets"],
                "sources": inferred["sources"],
                "selected": inferred["selected"],
                "remembered_mapping": STATE.mapping or {},
            })
            return

        try:
            df, report = clean(raw_df, column_map=mapping_body)
        except (MissingColumnsError, EmptyDatasetError) as e:
            return self._send_json({"error": str(e)}, 400)
        del raw_df

        index = build_index(df)
        with STATE.lock:
            STATE.mapping = mapping_body or (inferred["selected"] if inferred else STATE.mapping)

        with STATE.lock:
            STATE.df = df
            STATE.index = index
            STATE.filename = filename

        self._send_json({
            "filename": filename,
            "rows_in": report.rows_in,
            "rows_out": report.rows_out,
            "duplicates_removed": report.duplicates_removed,
            "currency_stripped": report.currency_stripped,
            "dates_failed": report.dates_failed,
            "blanks_filled": report.blanks_filled,
            "total_mismatches": report.total_mismatches[:5],
            "total_mismatches_count": len(report.total_mismatches),
            "mapping": STATE.mapping or {},
        })

    def _handle_ask(self):
        """Fast half of asking a question: routing + pandas only. classify()
        is a model call (small — a single-label prediction) so it goes
        through MODEL_LOCK, but it's short; answer() itself is pure pandas
        and never touches the model. Returns as soon as the Finding exists,
        so the receipt can be on screen well before narration is ready."""
        body = self._read_json()
        question = (body.get("question") or "").strip()
        if not question:
            return self._send_json({"error": "empty question"}, 400)

        with STATE.lock:
            df, index = STATE.df, STATE.index
        if df is None:
            return self._send_json({"error": "no file loaded yet"}, 400)

        try:
            with MODEL_LOCK:
                label = classify(question)
        except QuestionTooLongError as e:
            return self._send_json({"error": str(e)}, 400)
        route = label if label in HANDLERS else None
        finding = answer(df, question, index, label=label)
        self._send_json(_finding_to_json(finding, question, route))

    def _handle_narrate(self):
        """Slow half: recomputes the same Finding (cheap — pandas, not the
        model) and calls the real explain() — including its verify-and-
        retry pass — for the narration. Takes the route/label /api/ask
        already resolved (`label` in the body) so this never reclassifies
        the question itself; that keeps this endpoint's only model call to
        the one that matters (generation) and avoids a second classify()
        call racing the first one for MODEL_LOCK. Kept as a separate
        request from /api/ask on purpose: the frontend already has a
        Finding to render before this one resolves, rather than blocking
        the whole receipt on 15+ seconds of generation."""
        body = self._read_json()
        question = (body.get("question") or "").strip()
        if not question:
            return self._send_json({"error": "empty question"}, 400)

        with STATE.lock:
            df, index = STATE.df, STATE.index
        if df is None:
            return self._send_json({"error": "no file loaded yet"}, 400)

        label = body.get("label")  # None is valid: "no handler matched" case
        finding = answer(df, question, index, label=label)
        try:
            with MODEL_LOCK:
                narration = explain(finding, question)
        except QuestionTooLongError as e:
            return self._send_json({"error": str(e)}, 400)

        self._send_json({
            "text": narration.text,
            "verified": narration.verified,
            "unsupported": narration.unsupported,
            "retried": narration.retried,
        })

    def _handle_export(self):
        body = self._read_json()
        kind = body.get("kind")
        date = _today_iso()

        if kind == "answer":
            question = (body.get("question") or "").strip()
            finding_data = body.get("finding") or {}
            report_lines = _finding_lines(
                question,
                type("FindingView", (), {
                    "headline": finding_data.get("headline", ""),
                    "facts": finding_data.get("facts") or {},
                    "table": pd.DataFrame(finding_data.get("table") or []),
                })(),
                body.get("narration"),
            )
            slug = _slugify(question)
            filename = f"ledgerit-{slug}-{date}.pdf"
            out = _export_root() / filename
            out.write_bytes(_render_pdf(report_lines, f"Answer: {question}"))
        elif kind == "cleaning":
            report = body.get("report") or {}
            filename = (body.get("filename") or "cleaning-report").strip()
            slug = _slugify(filename.rsplit(".", 1)[0])
            out_name = f"ledgerit-{slug}-{date}.pdf"
            out = _export_root() / out_name
            out.write_bytes(_report_lines(filename, report, "Cleaning report", _cleaning_lines(report)))
        else:
            return self._send_json({"error": "unknown export type"}, 400)

        self._send_json({"path": str(out)})


def create_server(host: str, port: int) -> http.server.ThreadingHTTPServer:
    return http.server.ThreadingHTTPServer((host, port), Handler)
